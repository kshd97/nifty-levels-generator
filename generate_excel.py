
import pandas as pd
import numpy as np
import re
import os
import io

def get_day_sheets(xl):
    # Regex to match day names like tue, wed, Thu, fri, mon followed by numbers or nothing
    # Excluding 'Total', 'Max'
    sheet_names = xl.sheet_names
    day_pattern = re.compile(r'^(tue|wed|thu|fri|mon|sun|sat)', re.IGNORECASE)
    
    day_sheets = []
    for sheet in sheet_names:
        if day_pattern.match(sheet) and sheet.lower() not in ['total', 'max']:
            day_sheets.append(sheet)
            
    # Sort sheets? Ideally order matters (Tue, Wed, Thu...) if we do cumulative.
    # For now, let's trust the order in the file or sort manually if needed.
    # But names like 'tue6', 'wed6' might not sort alphabetically correctly vs 'thu6'.
    # Let's rely on list order for now or specific order if user provides.
    # Best guess: use the order they appear in the file.
    return day_sheets

def process_sheet_data(xl, sheet_name):
    try:
        # Dynamic header finding (reuse logic)
        # pd.read_excel supports ExcelFile object
        df_preview = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=10)
        header_row = -1
        for i, row in df_preview.iterrows():
            row_str = row.astype(str).str.lower().tolist()
            if 'strike' in row_str or 'chg in oi value' in row_str:
                header_row = i
                break
        
        if header_row == -1:
            return None

        df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row)
        df.columns = df.columns.astype(str).str.strip()
        
        # Clean columns
        df['Strike'] = pd.to_numeric(df['Strike'], errors='coerce')
        if 'Chg in OI Value' in df.columns:
            df['Call_Chg_OI_Val'] = pd.to_numeric(df['Chg in OI Value'], errors='coerce')
        else:
            return None # Essential column missing
            
        df['Call_VWAP'] = pd.to_numeric(df.get('VWAP', 0), errors='coerce')
        df['Call_LTP'] = pd.to_numeric(df.get('LTP (Chg %)', 0), errors='coerce')
        
        # Rename duplicate columns for Put side (usually .1 suffix)
        cols = df.columns.tolist()
        if 'Chg in OI Value.1' in cols:
            df['Put_Chg_OI_Val'] = pd.to_numeric(df['Chg in OI Value.1'], errors='coerce')
        if 'VWAP.1' in cols:
            df['Put_VWAP'] = pd.to_numeric(df['VWAP.1'], errors='coerce')
        if 'LTP (Chg %).1' in cols:
            df['Put_LTP'] = pd.to_numeric(df['LTP (Chg %).1'], errors='coerce')
            
        df = df.dropna(subset=['Strike'])
        # Ensure unique strikes to prevent reindexing errors
        df = df.drop_duplicates(subset=['Strike'])
        cols_to_fill = ['Call_Chg_OI_Val', 'Put_Chg_OI_Val', 'Call_VWAP', 'Put_VWAP', 'Call_LTP', 'Put_LTP']
        for c in cols_to_fill:
            if c in df.columns:
                df[c] = df[c].fillna(0)
            else:
                df[c] = 0
                
        return df[['Strike', 'Call_Chg_OI_Val', 'Call_VWAP', 'Call_LTP', 'Put_Chg_OI_Val', 'Put_VWAP', 'Put_LTP']]
    except Exception as e:
        print(f"Error reading sheet {sheet_name}: {e}")
        return None

def process_excel_file(input_source):
    """
    input_source: Can be a file path (str) or file-like object (bytes).
    Returns: Bytes of the modified Excel file.
    """
    
    # Load into buffer to allow in-memory modification
    if isinstance(input_source, str):
        with open(input_source, 'rb') as f:
            buffer = io.BytesIO(f.read())
    else:
         # Assume bytes or file-like
         if hasattr(input_source, 'read'):
             input_source.seek(0)
             buffer = io.BytesIO(input_source.read())
         else:
             buffer = io.BytesIO(input_source)
             
    # Create ExcelFile object for reading
    try:
        xl = pd.ExcelFile(buffer)
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return None
        
    day_sheets = get_day_sheets(xl)
    print(f"Found day sheets: {day_sheets}")
    
    all_data = []
    
    # 1. Process all daily sheets
    for idx, sheet in enumerate(day_sheets):
        df = process_sheet_data(xl, sheet)
        if df is not None:
            df['Sheet_Name'] = sheet
            df['Sheet_Index'] = idx
            all_data.append(df)
            
    if not all_data:
        print("No valid data.")
        return

    combined = pd.concat(all_data)
    
    # 2. Consistent Pricing Logic & Metrics Calculation
    # We need to compute Ref_Price for EACH day for the 'Total' sheet
    
    # Group by Strike to determine if we force LTP
    grouped = combined.groupby('Strike')
    
    strike_pricing_mode = {} # Strike -> 'LTP' or 'Standard'
    
    for strike, group in grouped:
        group = group.sort_values('Sheet_Index')
        first_day = group.iloc[0]
        if first_day['Call_VWAP'] <= 0:
            strike_pricing_mode[strike] = 'LTP'
        else:
            strike_pricing_mode[strike] = 'Standard'
            
    # Apply pricing
    def get_ref_price(row, side='Call'):
        strike = row['Strike']
        mode = strike_pricing_mode.get(strike, 'Standard')
        
        vwap = row[f'{side}_VWAP']
        ltp = row[f'{side}_LTP']
        
        if mode == 'LTP':
            return ltp
        else:
            return vwap if vwap > 0 else ltp

    combined['Call_Ref_Price'] = combined.apply(lambda r: get_ref_price(r, 'Call'), axis=1)
    combined['Put_Ref_Price'] = combined.apply(lambda r: get_ref_price(r, 'Put'), axis=1)
    
    combined['Call_BEP'] = combined['Strike'] + combined['Call_Ref_Price']
    combined['Put_BEP'] = combined['Strike'] - combined['Put_Ref_Price']
    
    # 3. Build 'Total' Sheet Data with MultiIndex Headers
    # Structure: (Day, Metric)
    # LOGIC UPDATE:
    # First Day (e.g. Tue): Single values.
    # Subsequent Days (e.g. Wed): Cumulative Sum (Money) and Cumulative Avg (Ref Price) of ALL days up to that point.
    
    # Base DataFrame with Strike
    dfs_to_concat = []
    base_df = pd.DataFrame({'Strike': sorted(combined['Strike'].unique())}).set_index('Strike')
    
    # Initialize Cumulative Trackers
    cum_ce_money = pd.Series(0, index=base_df.index)
    cum_pe_money = pd.Series(0, index=base_df.index)
    
    # Track Ref Prices for Cumulative Average
    # We need to store all individual ref prices to compute running average
    # Dictionary mapping Strike -> List of Ref Prices seen so far
    # Or cleaner: Since we process days sequentially, we can maintain the running sum and count
    
    cum_ce_ref_sum = pd.Series(0.0, index=base_df.index)
    cum_ce_ref_count = pd.Series(0, index=base_df.index)
    
    cum_pe_ref_sum = pd.Series(0.0, index=base_df.index)
    cum_pe_ref_count = pd.Series(0, index=base_df.index)
    
    # Iterate through SORTED day sheets (Order matters for cumulative)
    # day_sheets are ['tue', 'wed', 'thu'...] based on extraction order. 
    # Assuming extraction order is correct or file has them in order.
    
    # We also need a way to pass these cumulative metrics to the Max sheet generator
    # So we'll store the calculated cumulative DFs in a dict
    daily_calculated_dfs = {} 
    
    for idx, sheet in enumerate(day_sheets):
        # Get daily data
        day_data = combined[combined['Sheet_Name'] == sheet].set_index('Strike')
        
        # Align to base index (fill missing strikes with 0 for money, NaN for ref)
        day_data = day_data.reindex(base_df.index)
        
        # 1. Update Cumulative Money
        # FillNa(0) for money addition
        current_ce_money = day_data['Call_Chg_OI_Val'].fillna(0)
        current_pe_money = day_data['Put_Chg_OI_Val'].fillna(0)
        
        cum_ce_money = cum_ce_money.add(current_ce_money, fill_value=0)
        cum_pe_money = cum_pe_money.add(current_pe_money, fill_value=0)
        
        # 2. Update Cumulative Ref Price (Running Average)
        # Ref price only exists if day_data has it.
        # Logic: If Day 1, Run Avg = Day 1 Ref.
        # If Day 2, Run Avg = Mean(Day 1 Ref, Day 2 Ref).
        # We need to handle if a strike is missing in a day.
        # "Use LTP if VWAP missing" logic is already in 'Call_Ref_Price'
        
        # We need to know if the strike "exists" for this day to include in average?
        # Or do we include 0? No, defined ref price.
        # If Strike is present in input, it has a Ref Price (or LTP).
        # If Strike is NOT present in input for this day, do we carry forward or ignore?
        # Usually average is over "active" days? Or all days?
        # Let's assume average over present data points.
        
        # Call Ref
        valid_ce_mask = day_data['Call_Ref_Price'].notna() & (day_data['Call_Ref_Price'] > 0)
        cum_ce_ref_sum = cum_ce_ref_sum.add(day_data['Call_Ref_Price'].fillna(0), fill_value=0)
        cum_ce_ref_count = cum_ce_ref_count.add(valid_ce_mask.astype(int), fill_value=0)
        
        # Put Ref
        valid_pe_mask = day_data['Put_Ref_Price'].notna() & (day_data['Put_Ref_Price'] > 0)
        cum_pe_ref_sum = cum_pe_ref_sum.add(day_data['Put_Ref_Price'].fillna(0), fill_value=0)
        cum_pe_ref_count = cum_pe_ref_count.add(valid_pe_mask.astype(int), fill_value=0)
        
        # Calculate Running Avg
        # Avoid division by zero
        avg_ce_ref = cum_ce_ref_sum.div(cum_ce_ref_count).replace([np.inf, -np.inf], 0).fillna(0)
        avg_pe_ref = cum_pe_ref_sum.div(cum_pe_ref_count).replace([np.inf, -np.inf], 0).fillna(0)
        
        # 3. Calculate BEP based on Cumulative Avg Ref
        ce_bep = base_df.index + avg_ce_ref
        pe_bep = base_df.index - avg_pe_ref
        
        # 4. Construct DataFrame for this Day
        # Columns: CE BEP, CE Money (Cumulative), PE Money (Cumulative), PE BEP
        
        subset = pd.DataFrame(index=base_df.index)
        subset['CE BEP'] = ce_bep
        subset['CE Money'] = cum_ce_money
        subset['PE Money'] = cum_pe_money
        subset['PE BEP'] = pe_bep
        
        # Store for Max Sheet usage
        # We need to store the Avg Ref as well for Max Sheet levels
        subset['Avg CE Ref'] = avg_ce_ref
        subset['Avg PE Ref'] = avg_pe_ref
        daily_calculated_dfs[sheet] = subset.copy()
        
        # Drop aux cols for Total Sheet display
        display_subset = subset[['CE BEP', 'CE Money', 'PE Money', 'PE BEP']]
        
        # Create MultiIndex: (SheetName, Metric)
        # Note: Columns are effectively "Cumulative up to SheetName"
        columns = pd.MultiIndex.from_product([[sheet], display_subset.columns])
        display_subset.columns = columns
        
        dfs_to_concat.append(display_subset)
        
        # Add Gap Column
        gap_df = pd.DataFrame(index=base_df.index, columns=pd.MultiIndex.from_tuples([('', '')]))
        dfs_to_concat.append(gap_df)
        
    # Remove last gap
    if dfs_to_concat:
        dfs_to_concat.pop()
        
    # Concatenate all daily blocks
    final_total_df = pd.concat([base_df] + dfs_to_concat, axis=1)
    
    # Force MultiIndex conversion
    if isinstance(final_total_df.columns, pd.Index) and not isinstance(final_total_df.columns, pd.MultiIndex):
        final_total_df.columns = pd.MultiIndex.from_tuples(final_total_df.columns)
    
    # 5. Build 'Max' Sheet Data (Day-by-Day Summary)
    # Using the CUMULATIVE data we just calculated
    
    max_dfs = []
    
    # Helper for Max Summary
    def get_top5_df(df_source, ce_metric, pe_metric, label):
        # Metrics are already in df_source (which is the cumulative subset for that day)
        
        # Sort Calls (Top 5 by OI - Cumulative)
        top_calls = df_source.sort_values(ce_metric, ascending=False).head(5)
        # Sort Puts (Top 5 by OI - Cumulative)
        top_puts = df_source.sort_values(pe_metric, ascending=False).head(5)
        
        # Re-sort by Strike
        top_calls = top_calls.sort_index(ascending=False)
        top_puts = top_puts.sort_index(ascending=False)
        
        # Re-sort for display (Ease of viewing)
        # However, the user's image shows R1, R2, R3, R4, R5 tags next to the rows.
        # And the rows seem to be sorted by Strike Descending? 
        # Image 2 shows: 25700, 25800, 25600, 25750, 25650... NOT sorted by Strike. 
        # It seems they are sorted by ... Rank? 
        # Wait, the user said "Can you make it look cleaner (image 2)". 
        # Image 2 has 25000, 25200, 25100... They are NOT sorted by Strike.
        # They seem to be sorted by Rank (OI Value).
        # "R1" is 25000 (Money 81). "R4" is 25200 (Money 44). "R2" is 25100 (32? No wait).
        # Actually, let's look at the ranks in Image 2.
        # 25000 -> R1. 25600 -> 1.
        # The user's image 2 shows a mix.
        
        # User RECENT instructions: "Is it possible to create separate tables... As of now it seems very crammed up".
        # User Previous instruction: "Levels are sorted by Strike Price... to visualize the range easily".
        
        # I will follow the "Cleaner" request (Image 2) layout structure.
        # Layout: [CE Table] [Gap] [Gap] [PE Table]
        # CE Table Columns: CE Strike | Money | AVWAP | BEP
        # PE Table Columns: PE Strike | Money | AVWAP | BEP
        
        # Sorting: The user's previous request was specific about Strike Sorting. 
        # I will keep Strike Sorting (Descending) as it's cleaner than random rank order. 
        # Unless user explicitly asks to revert to Rank order. 
        # The image 2 might just be an example of structure.
        
        # Calculate Sums
        ce_money_sum = top_calls[ce_metric].sum()
        pe_money_sum = top_puts[pe_metric].sum()
        
        res_data = []
        for strike in top_calls.index:
            val = top_calls.loc[strike, ce_metric]
            ref = top_calls.loc[strike, 'Avg CE Ref'] 
            res_data.append([strike, val, ref, strike+ref])
            
        sup_data = []
        for strike in top_puts.index:
            val = top_puts.loc[strike, pe_metric]
            ref = top_puts.loc[strike, 'Avg PE Ref']
            sup_data.append([strike, val, ref, strike-ref])
            
        # Ensure 5 rows
        while len(res_data) < 5: res_data.append([np.nan]*4)
        while len(sup_data) < 5: sup_data.append([np.nan]*4)
        
        # Add a Total Row (Index 6)
        res_data.append([np.nan, ce_money_sum, np.nan, np.nan])
        sup_data.append([np.nan, pe_money_sum, np.nan, np.nan])
        
        # Columns based on User Request: CE Strike | Money | AVWAP | CE BEP
        res_df = pd.DataFrame(res_data, columns=['CE Strike', 'Money', 'AVWAP', 'CE BEP'])
        sup_df = pd.DataFrame(sup_data, columns=['PE Strike', 'Money', 'AVWAP', 'PE BEP'])
        
        # Columns based on User Request: CE Strike | Money | AVWAP | CE BEP
        res_df = pd.DataFrame(res_data, columns=['CE Strike', 'Money', 'AVWAP', 'CE BEP'])
        sup_df = pd.DataFrame(sup_data, columns=['PE Strike', 'Money', 'AVWAP', 'PE BEP'])
        
        # Add 1 blank column between CE and PE (User Request: 1 col between CE/PE)
        gap_inner = pd.DataFrame(index=range(len(res_df)), columns=['']) # 1 col
        
        block = pd.concat([res_df, gap_inner, sup_df], axis=1)
        
        # Restore MultiIndex Header
        block.columns = pd.MultiIndex.from_product([[label], block.columns])
        return block

    # Generate Max blocks 
    for sheet in day_sheets:
        if sheet in daily_calculated_dfs:
            stats = daily_calculated_dfs[sheet]
            # pass 'CE Money', 'PE Money' as they are the column names in stats
            block = get_top5_df(stats, 'CE Money', 'PE Money', sheet)
            max_dfs.append(block)
            
            # Add gap between DAYS (User Request: 2 cols between days)
            gap = pd.DataFrame(index=range(7), columns=pd.MultiIndex.from_tuples([('', ''), (' ', ' ')])) # 2 cols
            max_dfs.append(gap)
            
    # Remove last gap
    if max_dfs:
        max_dfs.pop()
    
    final_max_df = pd.concat(max_dfs, axis=1)
    
    # Reset index to blank strings to avoid ugly numbers, and we'll hide the index column
    final_max_df.index = [''] * len(final_max_df)
    
    # 6. Write to Excel and Style
    print("Writing to Excel...")
    
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.utils import get_column_letter
    
    # We use the buffer for writing. 'mode=a' requires an existing file/buffer content.
    # buffer already has the original file content.
    buffer.seek(0)
    
    with pd.ExcelWriter(buffer, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        final_total_df.to_excel(writer, sheet_name='Total') 
        final_max_df.to_excel(writer, sheet_name='Max', index=True)
        
        # --- Styling Max Sheet ---
        ws_max = writer.sheets['Max']
        ws_max.column_dimensions['A'].hidden = True # Hide Index
        
        # Styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        no_border = Border()
        
        # Max Sheet Logic
        max_start_row = 3
        max_end_row = 8
        current_col = 2
        
        for item in max_dfs:
            # Check Gap
            is_gap = False
            if isinstance(item.columns, pd.MultiIndex):
                if str(item.columns[0][0]).strip() == '':
                     is_gap = True
            
            width = len(item.columns)
            
            if is_gap:
                # Clear Borders for Gap Columns
                for c_offset in range(width):
                    col_idx = current_col + c_offset
                    column_letter = get_column_letter(col_idx)
                    # Clear header rows
                    ws_max.cell(row=1, column=col_idx).border = no_border
                    ws_max.cell(row=2, column=col_idx).border = no_border
                    # Clear data rows
                    for r in range(max_start_row, max_end_row + 1):
                        ws_max.cell(row=r, column=col_idx).border = no_border
            else:
                # Data Block
                ce_cols = range(current_col, current_col + 4)
                gap_col = current_col + 4
                pe_cols = range(current_col + 5, current_col + 9)
                
                def apply_box_grid(ws, rows, cols, border):
                    for r in rows:
                        for c in cols:
                            ws.cell(row=r, column=c).border = border
                            
                # Apply Borders to CE Table (Header + Data)
                apply_box_grid(ws_max, [2], ce_cols, thin_border) # Header
                apply_box_grid(ws_max, range(max_start_row, max_end_row + 1), ce_cols, thin_border)
                
                # Apply Borders to PE Table (Header + Data)
                apply_box_grid(ws_max, [2], pe_cols, thin_border) # Header
                apply_box_grid(ws_max, range(max_start_row, max_end_row + 1), pe_cols, thin_border)

                # Ensure Inner Gap is Clean
                ws_max.cell(row=1, column=gap_col).border = no_border
                ws_max.cell(row=2, column=gap_col).border = no_border
                for r in range(max_start_row, max_end_row + 1):
                    ws_max.cell(row=r, column=gap_col).border = no_border
                    
            current_col += width

        # --- Styling Total Sheet ---
        ws_total = writer.sheets['Total']
        
        # Col index mapping: DataFrame col i -> Excel col i + 2 (A is Index)
        for i, col_tuple in enumerate(final_total_df.columns):
            if str(col_tuple[0]).strip() == '':
                excel_col_idx = i + 2
                # Clear borders for this column
                # Rows: 1 to len(final_total_df) + 2 (Header)
                total_rows = len(final_total_df) + 2
                for r in range(1, total_rows + 1):
                    ws_total.cell(row=r, column=excel_col_idx).border = no_border
                    
        # Auto-fit Columns for BOTH sheets
        for ws in [ws_max, ws_total]:
            for col in ws.columns:
                max_length = 0
                # Get column letter from the first cell
                col_idx = col[0].column
                column_letter = get_column_letter(col_idx)
                
                if ws == ws_max and column_letter == 'A': continue # Skip hidden index
                
                for cell in col:
                    try:
                        val = cell.value
                        if val:
                            if len(str(val)) > max_length:
                                max_length = len(str(val))
                    except:
                        pass
                
                # Add a bit of padding
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

    # Return the modified bytes
    return buffer.getvalue()

if __name__ == "__main__":
    file_path = 'Nifty 10th Feb expiry - 1.xlsx'
    if os.path.exists(file_path):
        new_bytes = process_excel_file(file_path)
        if new_bytes:
            with open(file_path, 'wb') as f:
                f.write(new_bytes)
            print(f"Successfully processed {file_path}")
    else:
        print(f"File not found: {file_path}")
